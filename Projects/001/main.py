import openpyxl
inv_file=openpyxl.load_workbook("Inventory.xlsx")
product_list= inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_suplier = {}
product_under_10_inv = {}

# print (product_list.max_row)
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

#Calculation of products per supplier
    if supplier_name in products_per_supplier:
        current_num_product = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_product + 1
    else:
        products_per_supplier[supplier_name] = 1



# Total value of inventory per supplier
    if supplier_name in total_value_per_suplier:
        current_total_value = total_value_per_suplier.get(supplier_name)
        total_value_per_suplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_suplier[supplier_name] = inventory * price

# products with inventory less than 10
    if inventory < 10:
        product_under_10_inv[product_num] = inventory

# Add value for total inventory price

    inventory_price.value = inventory * price 

print(products_per_supplier)
print(total_value_per_suplier)
print(product_under_10_inv)

inv_file.save("inventory_with_total_value.xlsx")

